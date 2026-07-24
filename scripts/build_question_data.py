from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path("/Users/zhaoxuan/Desktop/北森图形推理1-2_逐题分类与考点难度台账.xlsx")
OUTPUT = ROOT / "app/questions.json"

BANK1_ANSWERS = (
    "DBADBAACACCDABAAADDBDDCCBCBCBCBBCBACDDDDAABCBBAAADABDADABABCDADC"
    "ADDDAACADAAECCCDCCEADAACBDADDCADAAA"
)

# Confirmed duplicates inside Bank 2. These later copies are omitted.
BANK2_DUPLICATES = {
    44,
    51,
    97,
    110,
    115,
    118,
    119,
    121,
    140,
    155,
    160,
    162,
    163,
    166,
    169,
    170,
    171,
    173,
    174,
    175,
    176,
    179,
    181,
    182,
    183,
    184,
    188,
    190,
    192,
    194,
    197,
    199,
    200,
    201,
    204,
    205,
    210,
    211,
    219,
    221,
    231,
    234,
}

# When the same question occurs in both banks, retain the clearer Bank 2 copy.
BANK2_TO_BANK1 = {
    2: 2,
    3: 72,
    12: 4,
    13: 93,
    14: 8,
    23: 22,
    24: 9,
    30: 12,
    31: 7,
    32: 47,
    48: 52,
    56: 29,
    63: 78,
    65: 28,
    79: 99,
    85: 36,
    87: 85,
    92: 38,
    93: 63,
    95: 71,
    101: 34,
    109: 35,
    111: 37,
    117: 21,
    126: 15,
    149: 31,
    156: 48,
    158: 89,
    159: 23,
    164: 11,
    165: 45,
    167: 55,
    172: 16,
    177: 87,
    187: 65,
    196: 42,
    198: 94,
    214: 74,
    217: 54,
    218: 5,
    220: 17,
    224: 33,
    227: 46,
    228: 70,
    230: 58,
    233: 6,
    236: 10,
}

ANSWER_OVERRIDES = {
    64: "A",
    150: "B",
    151: "B",
    152: "A",
    185: "C",
    208: "C",
    213: "B",
    214: "A",
    220: "A",
    222: "D",
}

# Question 64 is the clear copy of the same item described at row 184.
METADATA_SOURCE_OVERRIDES = {64: 184}

BANK1_OPTION_COUNTS = {57: 3, 76: 5, 77: 5, 83: 5, 89: 5}

MANUAL_ANALYSES = {
    185: (
        "四个位置按同一置换关系变化：左上移到左下，右上移到左上，"
        "左下移到右上，右下保持不变。代入第三组后应为 C。"
    ),
    222: (
        "三角构件的大小与朝向交替变化，同时依次改变附着方位；延续该"
        "轮换关系，下一图对应 D。此题原题未标答案，按图形规律补判。"
    ),
}


def clean(value, fallback=""):
    if value is None:
        return fallback
    text = " ".join(str(value).replace("\n", " ").split())
    return text if text and text != "—" else fallback


def load_rows():
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=True)
    sheet = workbook["逐题台账"]
    headers = [cell.value for cell in sheet[4]]
    rows = {}
    for values in sheet.iter_rows(min_row=5, values_only=True):
        if not values[0]:
            continue
        row = dict(zip(headers, values))
        rows[str(row["题目ID"])] = row
    return rows


def explanation(row, source_id, answer):
    point = clean(row.get("一级考点"), "特殊规律")
    second = clean(row.get("二级考点"), "综合判断")
    third = clean(row.get("三级考点"), second)
    rule = clean(
        row.get("具体规律"),
        f"围绕“{second}”比较题干中各元素的变化，并用选项逐一验证。",
    )
    number = int(source_id.split("-")[1])
    if source_id.startswith("2-") and number in MANUAL_ANALYSES:
        rule = MANUAL_ANALYSES[number]
    analysis = rule
    if answer and f"{answer}" not in analysis[-12:]:
        analysis = f"{analysis} 因此选择 {answer}。"
    method = (
        f"题型：{clean(row.get('题型形式'), '图形推理')}。"
        f"先识别“{second}”，再核对“{third}”，最后用其余元素排除干扰项。"
    )
    return point, second, third, analysis, method


def main():
    rows = load_rows()
    manifest = {
        item["question"]: item
        for item in json.loads(
            (ROOT / "public/questions/beisen-2/manifest.json").read_text(encoding="utf-8")
        )
    }
    questions = []
    removed_bank1 = set(BANK2_TO_BANK1.values())

    for number in range(1, 100):
        if number in removed_bank1:
            continue
        source_id = f"1-{number}"
        row = rows[source_id]
        answer = BANK1_ANSWERS[number - 1]
        point, second, third, analysis, method = explanation(row, source_id, answer)
        questions.append(
            {
                "sourceId": source_id,
                "image": f"/questions/beisen-1/q{number:03d}.webp",
                "optionImages": [],
                "answer": answer,
                "optionCount": BANK1_OPTION_COUNTS.get(number, 4),
                "point": point,
                "difficulty": clean(row.get("难度"), "提高"),
                "finePoints": [second, third],
                "analysis": analysis,
                "method": method,
                "source": "题库1",
                "originalNumber": number,
            }
        )

    for number in range(1, 238):
        if number in BANK2_DUPLICATES:
            continue
        source_id = f"2-{number}"
        metadata_number = METADATA_SOURCE_OVERRIDES.get(number, number)
        row = rows[f"2-{metadata_number}"]
        answer = ANSWER_OVERRIDES.get(number, clean(rows[source_id].get("答案")))
        if answer not in "ABCDE":
            raise RuntimeError(f"Missing answer for {source_id}")
        asset = manifest[number]
        option_images = asset["optionImages"]
        option_count = len(option_images)
        if not option_count:
            bank1_match = BANK2_TO_BANK1.get(number)
            option_count = (
                BANK1_OPTION_COUNTS.get(bank1_match, 4)
                if bank1_match
                else (5 if answer == "E" else 4)
            )
        point, second, third, analysis, method = explanation(row, source_id, answer)
        questions.append(
            {
                "sourceId": source_id,
                "image": asset["image"],
                "optionImages": option_images,
                "answer": answer,
                "optionCount": option_count,
                "point": point,
                "difficulty": clean(row.get("难度"), "提高"),
                "finePoints": [second, third],
                "analysis": analysis,
                "method": method,
                "source": "题库2",
                "originalNumber": number,
            }
        )

    if len(questions) != 247:
        raise RuntimeError(f"Expected 247 unique questions, got {len(questions)}")
    if len({question["sourceId"] for question in questions}) != len(questions):
        raise RuntimeError("Duplicate source IDs")

    OUTPUT.write_text(
        json.dumps(questions, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "questions": len(questions),
                "bank1": sum(item["source"] == "题库1" for item in questions),
                "bank2": sum(item["source"] == "题库2" for item in questions),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
